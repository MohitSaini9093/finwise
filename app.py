from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openai
import os
from dotenv import load_dotenv
from tabula.io import read_pdf as tabula_read_pdf
import pandas as pd
import numpy as np
import re
from datetime import datetime
from werkzeug.utils import secure_filename
import tempfile
from fuzzywuzzy import fuzz
from fuzzywuzzy import process
from flask import after_this_request
import PyPDF2
import pytesseract
from pdf2image import convert_from_path
import cv2
import numpy as np
from PIL import Image
import io
import nltk
from nltk.tokenize import word_tokenize, sent_tokenize
from nltk.corpus import stopwords
from nltk.tag import pos_tag
from pdfminer.high_level import extract_text
from collections import defaultdict
import openpyxl

# Load environment variables
load_dotenv()

# Download required NLTK data
try:
    nltk.data.find('tokenizers/punkt')
    nltk.data.find('taggers/averaged_perceptron_tagger')
    nltk.data.find('corpora/stopwords')
except LookupError:
    nltk.download('punkt')
    nltk.download('averaged_perceptron_tagger')
    nltk.download('stopwords')

app = Flask(__name__)
CORS(app)  # Enable CORS for all routes
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOAD_FOLDER'] = tempfile.gettempdir()  # Use system temp directory

# Configure OpenAI API key from environment variable
openai.api_key = os.getenv('OPENAI_API_KEY')

# Configure Tesseract path - update this path according to your system
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'  # For Windows
# pytesseract.pytesseract.tesseract_cmd = '/usr/bin/tesseract'  # For Linux/Mac

# Column mapping configurations
COLUMN_PATTERNS = {
    'serial_no': [
        'serial no', 'sr no', 'sr.no', 'sno', 's.no', 'serial number'
    ],
    'transaction_date': [
        'transaction date', 'txn date', 'date', 'tran date', 'trans date'
    ],
    'value_date': [
        'value date', 'val date', 'value dt', 'effective date'
    ],
    'description': [
        'description', 'particulars', 'narration', 'details', 'transaction details',
        'remarks', 'transaction description'
    ],
    'cheque_no': [
        'cheque number', 'cheque no', 'chq no', 'cheque', 'instrument no',
        'instrument number', 'ref no', 'reference number'
    ],
    'debit': [
        'debit', 'withdrawal', 'dr', 'debit amount', 'withdrawal amt',
        'amount debited', 'withdrawals'
    ],
    'credit': [
        'credit', 'deposit', 'cr', 'credit amount', 'deposit amt',
        'amount credited', 'deposits'
    ],
    'balance': [
        'balance', 'closing balance', 'running balance', 'bal', 'closing bal',
        'available balance', 'balance amount'
    ]
}

def preprocess_image(image):
    """Preprocess image for better OCR results"""
    # Convert to grayscale
    gray = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2GRAY)
    
    # Apply thresholding to preprocess the image
    gray = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
    
    # Apply dilation to connect text components
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3,3))
    gray = cv2.dilate(gray, kernel, iterations=1)
    
    # Convert back to PIL Image
    return Image.fromarray(gray)

def extract_text_from_image(image):
    """Extract text from image using OCR"""
    try:
        # Preprocess the image
        processed_image = preprocess_image(image)
        
        # Perform OCR
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(processed_image, config=custom_config)
        
        return text
    except Exception as e:
        print(f"Error in OCR processing: {str(e)}")
        return ""

def extract_text_from_pdf(pdf_path, password=None):
    """Extract text from PDF using both text extraction and OCR"""
    try:
        # First try normal text extraction
        text = extract_text(pdf_path, password=password)
        
        # If no text found or very little text, try OCR
        if not text or len(text.strip()) < 100:
            print("Attempting OCR extraction...")
            try:
                # Convert PDF to images
                images = convert_from_path(pdf_path, password=password if password else None)
                
                # Extract text from each image
                text_parts = []
                for image in images:
                    text_parts.append(extract_text_from_image(image))
                
                text = "\n".join(text_parts)
            except Exception as e:
                print(f"Error in PDF to image conversion: {str(e)}")
        
        return text
    except Exception as e:
        print(f"Error extracting text from PDF: {str(e)}")
        return None

def extract_table_from_image(image):
    """Extract table data from image using OCR"""
    try:
        # Preprocess the image
        processed_image = preprocess_image(image)
        
        # Use Tesseract's table extraction
        custom_config = r'--oem 3 --psm 6 -c preserve_interword_spaces=1'
        data = pytesseract.image_to_data(processed_image, config=custom_config, output_type=pytesseract.Output.DATAFRAME)
        
        # Filter out empty text and low confidence results
        data = data[data.conf != -1]
        data = data[data.text.str.strip() != '']
        
        # Group text by lines
        lines = []
        current_line = []
        current_block = -1
        
        for _, row in data.iterrows():
            if row['block_num'] != current_block and current_line:
                lines.append(' '.join(current_line))
                current_line = []
            current_line.append(str(row['text']).strip())
            current_block = row['block_num']
        
        if current_line:
            lines.append(' '.join(current_line))
        
        return '\n'.join(lines)
    except Exception as e:
        print(f"Error in table extraction from image: {str(e)}")
        return ""

def analyze_text_content(text):
    """Analyze text content using NLTK for transaction identification"""
    try:
        # Tokenize text
        sentences = sent_tokenize(text)
        words = word_tokenize(text.lower())
        
        # Remove stopwords
        stop_words = set(stopwords.words('english'))
        words = [w for w in words if w not in stop_words]
        
        # POS tagging
        tagged = pos_tag(words)
        
        # Initialize scores
        scores = {
            'has_dates': 0,
            'has_amounts': 0,
            'has_transaction_keywords': 0,
            'has_bank_keywords': 0
        }
        
        # Check for dates using regex
        date_patterns = [
            r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b',
            r'\b\d{4}[-/]\d{1,2}[-/]\d{1,2}\b',
            r'\b\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}\b'
        ]
        for pattern in date_patterns:
            scores['has_dates'] += len(re.findall(pattern, text))
        
        # Check for amounts
        amount_patterns = [
            r'\$?\d+,?\d*\.\d{2}\b',
            r'(?:RS|INR|₹)\s*\d+,?\d*\.?\d*\b',
            r'\b\d+,?\d*\.?\d*\s*(?:CR|DR)\b'
        ]
        for pattern in amount_patterns:
            scores['has_amounts'] += len(re.findall(pattern, text))
        
        # Check for transaction keywords
        transaction_keywords = {
            'transfer', 'payment', 'deposit', 'withdrawal', 'credit', 'debit',
            'balance', 'transaction', 'upi', 'neft', 'rtgs', 'imps', 'atm'
        }
        scores['has_transaction_keywords'] = sum(1 for word in words if word in transaction_keywords)
        
        # Check for bank keywords
        bank_keywords = {
            'bank', 'sbi', 'hdfc', 'icici', 'axis', 'pnb', 'kotak', 'union', 'canara'
        }
        scores['has_bank_keywords'] = sum(1 for word in words if word in bank_keywords)
        
        # Calculate confidence score
        total_score = sum(scores.values())
        max_possible_score = 20  # Reasonable maximum score
        confidence = min(total_score / max_possible_score, 1.0)
        
        return confidence, scores
        
    except Exception as e:
        print(f"Error in text analysis: {str(e)}")
        return 0.0, {'has_dates': 0, 'has_amounts': 0, 'has_transaction_keywords': 0, 'has_bank_keywords': 0}

def get_transaction_category(description):
    """Determine transaction category based on description"""
    description = str(description).lower()
    
    # Define category patterns
    categories = {
        'Housing': [
            'rent', 'mortgage', 'house', 'property', 'maintenance', 'repair',
            'electricity', 'water', 'gas', 'utility', 'utilities'
        ],
        'Transportation': [
            'fuel', 'petrol', 'diesel', 'gas', 'parking', 'car', 'auto',
            'taxi', 'uber', 'ola', 'metro', 'bus', 'train', 'transport'
        ],
        'Food': [
            'restaurant', 'food', 'grocery', 'groceries', 'swiggy', 'zomato',
            'dining', 'cafe', 'coffee', 'lunch', 'dinner', 'breakfast'
        ],
        'Shopping': [
            'amazon', 'flipkart', 'myntra', 'shopping', 'mall', 'store',
            'clothing', 'shoes', 'accessories', 'electronics'
        ],
        'Entertainment': [
            'movie', 'theatre', 'netflix', 'amazon prime', 'hotstar',
            'entertainment', 'game', 'sports', 'music', 'spotify'
        ],
        'Healthcare': [
            'medical', 'medicine', 'doctor', 'hospital', 'clinic',
            'pharmacy', 'health', 'insurance', 'dental', 'fitness'
        ],
        'Education': [
            'school', 'college', 'university', 'course', 'training',
            'education', 'tuition', 'books', 'stationery'
        ],
        'Bills & Utilities': [
            'bill', 'recharge', 'mobile', 'phone', 'internet', 'wifi',
            'broadband', 'dth', 'cable', 'subscription'
        ],
        'Investment': [
            'investment', 'mutual fund', 'stock', 'share', 'bond',
            'dividend', 'interest', 'fd', 'fixed deposit'
        ],
        'Salary': [
            'salary', 'wage', 'pay', 'income', 'bonus', 'incentive'
        ]
    }
    
    # Check each category's patterns
    for category, patterns in categories.items():
        if any(pattern in description for pattern in patterns):
            return category
    
    return 'Others'  # Default category

def extract_bank_statement(pdf_path, password=None):
    """Extract transactions from PDF using improved extraction methods including OCR"""
    transactions = []
    account_details = {}
    
    try:
        # First try to extract text using combined method (text + OCR)
        text_content = extract_text_from_pdf(pdf_path, password)
        if not text_content:
            print("Warning: Failed to extract text from PDF")
        
        # Try to extract account details from the header
        if text_content:
            account_pattern = re.compile(r'Account\s+Name\s*:?\s*([^\n]+)', re.IGNORECASE)
            account_number_pattern = re.compile(r'Account\s+Number\s*:?\s*(\d+)', re.IGNORECASE)
            branch_pattern = re.compile(r'Branch\s+Name\s*:?\s*([^\n]+)', re.IGNORECASE)
            
            if account_match := account_pattern.search(text_content):
                account_details['account_name'] = account_match.group(1).strip()
            if number_match := account_number_pattern.search(text_content):
                account_details['account_number'] = number_match.group(1).strip()
            if branch_match := branch_pattern.search(text_content):
                account_details['branch'] = branch_match.group(1).strip()
        
        # Try tabula for table extraction
        tables = []
        try:
            # Try different table extraction methods
            extraction_methods = [
                {'lattice': True, 'stream': False},
                {'lattice': False, 'stream': True},
                {'lattice': False, 'stream': True, 'guess': True},
                {'lattice': False, 'stream': True, 'guess': False, 'relative_area': True}
            ]
            
            for method in extraction_methods:
                try:
                    current_tables = tabula_read_pdf(
                        pdf_path,
                        pages='all',
                        password=password,
                        multiple_tables=True,
                        pandas_options={'header': None},
                        **method
                    )
                    if current_tables:
                        tables.extend(current_tables)
                except Exception as e:
                    print(f"Table extraction method failed: {str(e)}")
                    continue
            
            # If no tables found, try OCR-based extraction
            if not tables:
                print("Attempting OCR-based table extraction...")
                images = convert_from_path(pdf_path, password=password if password else None)
                
                for image in images:
                    table_text = extract_table_from_image(image)
                    if table_text:
                        # Convert extracted text to DataFrame
                        lines = table_text.split('\n')
                        rows = [line.split() for line in lines]
                        if rows:
                            tables.append(pd.DataFrame(rows))
            
        except Exception as e:
            print(f"Table extraction error: {str(e)}")
            tables = []
        
        # Process tables if found
        if tables:
            for table in tables:
                if table.empty:
                    continue
                
                # Clean the table data
                table = table.replace({np.nan: ''})
                table = table.applymap(lambda x: str(x).strip() if isinstance(x, str) else x)
                
                # Try to identify headers
                for idx, row in table.iterrows():
                    headers = [str(val).strip().lower() for val in row.values]
                    
                    # Check if this is a header row by looking for key column names
                    if any(x in headers for x in ['date', 'description', 'debit', 'credit', 'balance', 'particulars', 'narration']):
                        column_mapping = identify_columns(headers)
                        
                        if not column_mapping:
                            continue
                        
                        # Process rows after header
                        for data_idx in range(idx + 1, len(table)):
                            try:
                                row_data = table.iloc[data_idx]
                                
                                # Skip rows that don't look like transactions
                                if all(pd.isna(val) or str(val).strip() == '' for val in row_data):
                                    continue
                                
                                # Extract transaction data with better error handling
                                transaction = extract_transaction_from_row(row_data, column_mapping)
                                if transaction:
                                    # Add account details if available
                                    if account_details:
                                        transaction.update({
                                            "ACCOUNT_NAME": account_details.get('account_name', ''),
                                            "ACCOUNT_NUMBER": account_details.get('account_number', ''),
                                            "BRANCH": account_details.get('branch', '')
                                        })
                                    transactions.append(transaction)
                                
                            except Exception as e:
                                print(f"Error processing row: {str(e)}")
                                continue
                        
                        # Break after processing the transaction table
                        if transactions:
                            break
        
        if not transactions:
            # If no transactions found through table extraction, try text-based extraction
            text_transactions = extract_transactions_from_text(text_content)
            if text_transactions:
                transactions.extend(text_transactions)
        
        if not transactions:
            raise Exception("No transactions found in the PDF. Please ensure the PDF contains transaction data in a readable format.")
        
        return transactions
        
    except Exception as e:
        raise Exception(f"Error processing PDF: {str(e)}")

def extract_transaction_from_row(row_data, column_mapping):
    """Extract transaction data from a row with better error handling"""
    try:
        # Extract serial number
        serial_no = None
        if 'serial_no' in column_mapping:
            serial_no = str(row_data[column_mapping['serial_no']]).strip()
            if not re.match(r'^\d+$', serial_no):
                serial_no = None
        
        # Extract dates
        tran_date = None
        value_date = None
        
        if 'transaction_date' in column_mapping:
            tran_date = parse_date(str(row_data[column_mapping['transaction_date']]))
        if 'value_date' in column_mapping:
            value_date = parse_date(str(row_data[column_mapping['value_date']]))
        
        # Extract description
        description = ""
        if 'description' in column_mapping:
            description = str(row_data[column_mapping['description']]).strip()
        
        # Skip if no valid description
        if not description or description.lower() in ['description', 'particulars', 'narration']:
            return None
        
        # Extract amounts
        debit_amount = 0
        credit_amount = 0
        balance = None
        
        if 'debit' in column_mapping:
            debit_str = str(row_data[column_mapping['debit']]).strip()
            if debit_str and debit_str not in ['-', 'na', 'nil']:
                debit_amount = parse_amount(debit_str)
        
        if 'credit' in column_mapping:
            credit_str = str(row_data[column_mapping['credit']]).strip()
            if credit_str and credit_str not in ['-', 'na', 'nil']:
                credit_amount = parse_amount(credit_str)
        
        if 'balance' in column_mapping:
            balance_str = str(row_data[column_mapping['balance']]).strip()
            if balance_str and balance_str not in ['-', 'na', 'nil']:
                balance = parse_amount(balance_str)
        
        # Get cheque number if available
        cheque_no = ""
        if 'cheque_no' in column_mapping:
            cheque_str = str(row_data[column_mapping['cheque_no']]).strip()
            if cheque_str and cheque_str not in ['-', 'na', 'nil']:
                cheque_no = cheque_str
        
        # Determine transaction type and category
        transaction_type = "DEBIT" if debit_amount > 0 else "CREDIT" if credit_amount > 0 else "UNKNOWN"
        category = get_transaction_category(description)
        
        # Create transaction entry
        return {
            "SERIAL_NO": serial_no or "",
            "DATE": tran_date or "",
            "VALUE_DATE": value_date or tran_date or "",
            "DESCRIPTION": description,
            "CHEQUE_NO": cheque_no,
            "DEBIT": abs(debit_amount),
            "CREDIT": abs(credit_amount),
            "BALANCE": balance or 0.0,
            "TYPE": transaction_type,
            "CATEGORY": category
        }
        
    except Exception as e:
        print(f"Error extracting transaction from row: {str(e)}")
        return None

def extract_transactions_from_text(text_content):
    """Extract transactions from text content when table extraction fails"""
    if not text_content:
        return []
    
    transactions = []
    lines = text_content.split('\n')
    
    # Common date patterns
    date_pattern = r'\b\d{1,2}[-/]\d{1,2}[-/]\d{2,4}\b'
    amount_pattern = r'(?:(?:RS|INR|₹)\s*)?(\d+(?:,\d+)*(?:\.\d{2})?)'
    
    current_transaction = {}
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Look for date
        if re.search(date_pattern, line):
            # If we have a previous transaction, save it
            if current_transaction:
                transactions.append(current_transaction)
                current_transaction = {}
            
            # Start new transaction
            date_match = re.search(date_pattern, line)
            current_transaction['DATE'] = parse_date(date_match.group(0))
            current_transaction['DESCRIPTION'] = line.replace(date_match.group(0), '').strip()
            
            # Look for amounts in the same line
            amounts = re.findall(amount_pattern, line)
            if amounts:
                if 'Dr' in line or 'DR' in line or 'WITHDRAWAL' in line.upper():
                    current_transaction['DEBIT'] = parse_amount(amounts[-1])
                    current_transaction['CREDIT'] = 0
                elif 'Cr' in line or 'CR' in line or 'DEPOSIT' in line.upper():
                    current_transaction['DEBIT'] = 0
                    current_transaction['CREDIT'] = parse_amount(amounts[-1])
        
        # If we're in a transaction and find amounts
        elif current_transaction and re.search(amount_pattern, line):
            amounts = re.findall(amount_pattern, line)
            if amounts:
                if 'Dr' in line or 'DR' in line or 'WITHDRAWAL' in line.upper():
                    current_transaction['DEBIT'] = parse_amount(amounts[-1])
                    current_transaction['CREDIT'] = 0
                elif 'Cr' in line or 'CR' in line or 'DEPOSIT' in line.upper():
                    current_transaction['DEBIT'] = 0
                    current_transaction['CREDIT'] = parse_amount(amounts[-1])
                
                # If we have both debit and credit, this transaction is complete
                if 'DEBIT' in current_transaction and 'CREDIT' in current_transaction:
                    transactions.append(current_transaction)
                    current_transaction = {}
    
    # Add the last transaction if any
    if current_transaction:
        transactions.append(current_transaction)
    
    # Clean and standardize transactions
    cleaned_transactions = []
    for trans in transactions:
        if 'DATE' in trans and ('DEBIT' in trans or 'CREDIT' in trans):
            cleaned_trans = {
                "SERIAL_NO": "",
                "DATE": trans.get('DATE', ''),
                "VALUE_DATE": trans.get('DATE', ''),
                "DESCRIPTION": trans.get('DESCRIPTION', ''),
                "CHEQUE_NO": "",
                "DEBIT": trans.get('DEBIT', 0),
                "CREDIT": trans.get('CREDIT', 0),
                "BALANCE": 0.0,
                "TYPE": "DEBIT" if trans.get('DEBIT', 0) > 0 else "CREDIT"
            }
            cleaned_transactions.append(cleaned_trans)
    
    return cleaned_transactions

def find_best_column_match(header, column_type):
    """Find the best matching column name using fuzzy string matching"""
    patterns = COLUMN_PATTERNS[column_type]
    best_match = process.extractOne(header.lower(), patterns)
    return best_match[1] if best_match and best_match[1] >= 75 else None

def identify_columns(headers):
    """Identify column types using fuzzy matching"""
    column_mapping = {}
    for idx, header in enumerate(headers):
        for col_type in COLUMN_PATTERNS.keys():
            if find_best_column_match(header, col_type):
                column_mapping[col_type] = idx
                break
    return column_mapping

def parse_amount(amount_str):
    """Parse amount string to float, handling different formats"""
    try:
        # Return 0 for empty, undefined, or NaN values
        if not amount_str or pd.isna(amount_str) or str(amount_str).lower() in ['undefined', 'nan', '-', 'na', 'nil']:
            return 0.0
        
        # Convert to string and clean up
        amount_str = str(amount_str).strip().upper()
        
        # Extract amount from currency format (e.g., "Debit ₹1000")
        amount_match = re.search(r'(?:DEBIT|CREDIT)?\s*(?:RS|INR|₹)?\s*(\d+(?:,\d+)*(?:\.\d{2})?)', amount_str)
        if amount_match:
            amount_str = amount_match.group(1)
        
        # Remove any remaining currency symbols, commas and other non-numeric characters
        amount_str = re.sub(r'[^\d.-]', '', amount_str)
        
        # Convert to float
        amount = float(amount_str)
        return abs(amount)  # Return absolute value
        
    except (ValueError, TypeError, AttributeError):
        return 0.0

def parse_date(date_str):
    """Parse date string in various formats"""
    try:
        # Return None for empty, undefined, or NaN values
        if not date_str or pd.isna(date_str) or str(date_str).lower() in ['undefined', 'nan']:
            return None
        
        # Clean up the date string
        date_str = str(date_str).strip()
        
        # Try parsing with different formats
        date_formats = [
            '%d/%m/%Y', '%d-%m-%Y',  # DD/MM/YYYY or DD-MM-YYYY
            '%Y/%m/%d', '%Y-%m-%d',  # YYYY/MM/DD or YYYY-MM-DD
            '%m/%d/%Y', '%m-%d-%Y',  # MM/DD/YYYY or MM-DD-YYYY
            '%d/%m/%y', '%d-%m-%y'   # DD/MM/YY or DD-MM-YY
        ]
        
        for fmt in date_formats:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                return date_obj.strftime('%d/%m/%Y')  # Return in standard format
            except ValueError:
                continue
        
        # If no format matches, try extracting date using regex
        date_patterns = [
            r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})',  # DD/MM/YYYY
            r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})',  # YYYY/MM/DD
            r'(\d{1,2})[/-](\d{1,2})[/-](\d{2})'   # DD/MM/YY
        ]
        
        for pattern in date_patterns:
            if match := re.search(pattern, date_str):
                try:
                    if pattern == date_patterns[1]:  # YYYY/MM/DD
                        return f"{match.group(2)}/{match.group(3)}/{match.group(1)}"
                    elif pattern == date_patterns[2]:  # DD/MM/YY
                        year = int(match.group(3))
                        year = year + 2000 if year < 50 else year + 1900
                        return f"{match.group(1)}/{match.group(2)}/{year}"
                    else:
                        return f"{match.group(1)}/{match.group(2)}/{match.group(3)}"
                except:
                    continue
        
        return None
        
    except (ValueError, TypeError, AttributeError):
        return None

def check_pdf_password_protected(pdf_path):
    """Check if PDF is password protected and return status"""
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            return pdf_reader.is_encrypted
    except Exception as e:
        print(f"Error checking PDF password protection: {str(e)}")
        return False

def get_fallback_recommendations(income, expenses, savings):
    """Provide fallback recommendations when AI service is unavailable"""
    total_expenses = sum(expenses) if expenses else 0
    savings_rate = (savings / income * 100) if income > 0 else 0
    
    recommendations = [
        "📈 Aim to save at least 20% of your monthly income",
        "🏦 Build an emergency fund covering 3-6 months of expenses",
        "💰 Consider diversifying your investments across different asset classes",
        "📊 Track your expenses regularly and categorize them",
        "💳 Try to minimize high-interest debt"
    ]
    
    if savings_rate < 20:
        recommendations.append("⚠️ Your current savings rate is below recommended levels. Consider reducing non-essential expenses.")
    if total_expenses > (income * 0.7):
        recommendations.append("⚠️ Your expenses are high relative to your income. Look for areas to cut back.")
    
    return "\n".join(recommendations)

def get_fallback_chat_response(user_message, context):
    """Provide fallback chat responses when AI service is unavailable"""
    message_lower = user_message.lower()
    
    # Basic response templates based on common questions
    if any(word in message_lower for word in ['save', 'saving', 'savings']):
        return "To improve your savings, try the 50/30/20 rule: 50% for needs, 30% for wants, and 20% for savings. Track your expenses regularly and look for areas where you can cut back."
    
    if any(word in message_lower for word in ['invest', 'investment']):
        return "For investments, consider diversifying across different assets like mutual funds, stocks, and fixed deposits. Start with low-risk options and gradually expand your portfolio based on your risk tolerance."
    
    if any(word in message_lower for word in ['budget', 'spending']):
        return "Create a detailed budget by tracking all expenses. Use categories to understand your spending patterns. Look for areas where you can reduce unnecessary expenses."
    
    if any(word in message_lower for word in ['debt', 'loan']):
        return "Prioritize paying off high-interest debt first. Consider the debt avalanche method (paying highest interest first) or debt snowball method (paying smallest debts first)."
    
    # Default response
    return "I can help you with financial advice about savings, investments, budgeting, and debt management. What specific aspect would you like to know more about?"

def clean_transaction_data(transaction):
    """Clean and validate transaction data"""
    try:
        # Basic data validation
        if all(not transaction.get(key) or str(transaction.get(key)).lower() in ['n/a', 'na', 'undefined', 'none', ''] 
               for key in ['DATE', 'DESCRIPTION', 'DEBIT', 'CREDIT']):
            return None

        # Clean and validate date
        date = str(transaction.get('DATE', '')).strip()
        if date.lower() in ['n/a', 'na', 'undefined', 'none', '']:
            date = None
        else:
            date = parse_date(date)

        # Clean and validate description
        description = str(transaction.get('DESCRIPTION', '')).strip()
        if description.lower() in ['n/a', 'na', 'undefined', 'none', '']:
            description = 'Unspecified Transaction'

        # Clean and validate amounts
        debit = parse_amount(transaction.get('DEBIT', 0))
        credit = parse_amount(transaction.get('CREDIT', 0))
        balance = parse_amount(transaction.get('BALANCE', 0))

        # Determine transaction type
        if debit > 0:
            amount = debit
            transaction_type = 'DEBIT'
        elif credit > 0:
            amount = credit
            transaction_type = 'CREDIT'
        else:
            amount = 0
            transaction_type = 'UNKNOWN'

        # Get or guess category
        category = str(transaction.get('CATEGORY', '')).strip()
        if not category or category.lower() in ['n/a', 'na', 'undefined', 'none', '']:
            category = get_transaction_category(description)

        # Create cleaned transaction
        cleaned_trans = {
            "DATE": date or '',
            "DESCRIPTION": description,
            "AMOUNT": amount,
            "TYPE": transaction_type,
            "CATEGORY": category,
            "DEBIT": debit,
            "CREDIT": credit,
            "BALANCE": balance,
            "ACCOUNT_NAME": str(transaction.get('ACCOUNT_NAME', '')).strip() or 'Unknown Account',
            "ACCOUNT_NUMBER": str(transaction.get('ACCOUNT_NUMBER', '')).strip() or '',
            "REFERENCE": str(transaction.get('CHEQUE_NO', '')).strip() or ''
        }

        return cleaned_trans if cleaned_trans["AMOUNT"] > 0 else None

    except Exception as e:
        print(f"Error cleaning transaction: {str(e)}")
        return None

@app.route('/api/chat', methods=['POST'])
def chat():
    try:
        # Validate request data
        data = request.json
        if not data:
            return jsonify({
                "success": False,
                "error": "No data provided in request"
            }), 400

        user_message = data.get('message')
        if not user_message:
            return jsonify({
                "success": False,
                "error": "No message provided"
            }), 400

        context = data.get('context', {})
        
        try:
            # Try OpenAI API first
            financial_context = f"""
            User Profile:
            - Name: {context.get('userProfile', {}).get('name', 'User')}
            - Monthly Income: ₹{context.get('monthlyIncome', 0)}
            - Recent Transactions: {len(context.get('transactions', []))} transactions
            
            User Question: {user_message}
            """
            
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a knowledgeable financial advisor. Provide clear, concise advice based on the user's financial situation."},
                    {"role": "user", "content": financial_context}
                ],
                max_tokens=150,
                temperature=0.7
            )
            
            ai_response = response.choices[0].message.content.strip()
            
        except Exception as api_error:
            print(f"OpenAI API Error: {str(api_error)}")
            # Use fallback response system
            ai_response = get_fallback_chat_response(user_message, context)
        
        return jsonify({
            "success": True,
            "response": ai_response
        })
        
    except Exception as e:
        print(f"Unexpected Error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "An unexpected error occurred"
        }), 500

@app.route('/api/financial-advice', methods=['POST'])
def financial_advice():
    try:
        # Validate request data
        data = request.json
        if not data:
            return jsonify({
                "success": False,
                "error": "No data provided in request"
            }), 400

        income = data.get('income', 0)
        expenses = data.get('expenses', [])
        savings = data.get('savings', 0)

        # Validate financial data
        if income <= 0:
            return jsonify({
                "success": False,
                "error": "Income must be greater than 0"
            }), 400
        
        try:
            # Try OpenAI API first
            financial_data = f"""
            Financial Overview:
            - Monthly Income: ₹{income}
            - Total Expenses: ₹{sum(expenses)}
            - Monthly Savings: ₹{savings}
            - Savings Rate: {(savings/income * 100):.1f}%
            
            Please provide personalized financial recommendations based on this data.
            """
            
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "You are a financial advisor. Provide actionable recommendations based on the user's financial data."},
                    {"role": "user", "content": financial_data}
                ],
                max_tokens=200,
                temperature=0.7
            )
            
            advice = response.choices[0].message.content.strip()
            
        except Exception as api_error:
            print(f"OpenAI API Error: {str(api_error)}")
            # Use fallback recommendations system
            advice = get_fallback_recommendations(income, expenses, savings)
        
        return jsonify({
            "success": True,
            "advice": advice
        })
        
    except Exception as e:
        print(f"Unexpected Error: {str(e)}")
        return jsonify({
            "success": False,
            "error": "An unexpected error occurred"
        }), 500

@app.route('/api/process-pdf', methods=['POST'])
def process_pdf():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Only PDF files are allowed'}), 400
        
        password = request.form.get('password')
        
        try:
            filename = secure_filename(f"{datetime.now().timestamp()}_{file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            try:
                if check_pdf_password_protected(filepath) and not password:
                    return jsonify({
                        'error': 'PASSWORD_REQUIRED',
                        'message': 'This PDF is password protected. Please provide the password.',
                        'requiresPassword': True,
                        'filename': filename
                    }), 401
                
                try:
                    raw_transactions = extract_bank_statement(filepath, password)
                    if not raw_transactions:
                        raise Exception("No transactions found in the PDF")
                    
                    # Clean and validate transactions
                    cleaned_transactions = []
                    for trans in raw_transactions:
                        cleaned_trans = clean_transaction_data(trans)
                        if cleaned_trans:
                            cleaned_transactions.append(cleaned_trans)
                    
                    if not cleaned_transactions:
                        raise Exception("No valid transactions found after cleaning")
                    
                    # Create DataFrame with cleaned data
                    df = pd.DataFrame(cleaned_transactions)
                    
                    # Format numeric columns
                    numeric_columns = ['AMOUNT', 'DEBIT', 'CREDIT', 'BALANCE']
                    for col in numeric_columns:
                        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).round(2)
                    
                    # Format date column
                    if 'DATE' in df.columns:
                        df['DATE'] = pd.to_datetime(df['DATE'], format='%d/%m/%Y', errors='coerce')
                        df['DATE'] = df['DATE'].dt.strftime('%d/%m/%Y').fillna('')
                    
                    # Generate filenames
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    base_filename = f"transactions_{timestamp}"
                    csv_filename = f"{base_filename}.csv"
                    excel_filename = f"{base_filename}.xlsx"
                    
                    csv_filepath = os.path.join(app.config['UPLOAD_FOLDER'], csv_filename)
                    excel_filepath = os.path.join(app.config['UPLOAD_FOLDER'], excel_filename)
                    
                    # Save as CSV with proper encoding
                    df.to_csv(csv_filepath, index=False, encoding='utf-8-sig')
                    
                    # Save as Excel with formatting
                    with pd.ExcelWriter(excel_filepath, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Transactions')
                        workbook = writer.book
                        worksheet = writer.sheets['Transactions']
                        
                        # Format headers
                        header_style = openpyxl.styles.NamedStyle(name='header_style')
                        header_style.font = openpyxl.styles.Font(bold=True)
                        header_style.fill = openpyxl.styles.PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
                        header_style.alignment = openpyxl.styles.Alignment(horizontal='center')
                        
                        for col in range(len(df.columns)):
                            cell = worksheet.cell(row=1, column=col+1)
                            cell.style = header_style
                        
                        # Format numeric columns
                        money_format = openpyxl.styles.numbers.BUILTIN_FORMATS[44]
                        for col in numeric_columns:
                            if col in df.columns:
                                col_letter = openpyxl.utils.get_column_letter(df.columns.get_loc(col) + 1)
                                for row in range(2, len(df) + 2):
                                    cell = worksheet.cell(row=row, column=df.columns.get_loc(col) + 1)
                                    cell.number_format = money_format
                        
                        # Auto-adjust column widths
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                    # Clean up the PDF file
                    if os.path.exists(filepath):
                        os.remove(filepath)
                    
                    return jsonify({
                        'success': True,
                        'transactions': cleaned_transactions,
                        'message': f'Successfully extracted {len(cleaned_transactions)} transactions',
                        'columns': list(df.columns),
                        'downloadReady': True,
                        'excelFilePath': excel_filepath,
                        'excelFileName': excel_filename
                    })
                    
                except Exception as e:
                    if str(e) == "PASSWORD_REQUIRED":
                        return jsonify({
                            'error': 'PASSWORD_REQUIRED',
                            'message': 'This PDF is password protected. Please provide the password.',
                            'requiresPassword': True,
                            'filename': filename
                        }), 401
                    raise
                
            except Exception as e:
                if os.path.exists(filepath):
                    os.remove(filepath)
                raise
                
        except Exception as e:
            print(f"File Handling Error: {str(e)}")
            return jsonify({
                'error': 'Error handling file',
                'message': str(e)
            }), 500
            
    except Exception as e:
        print(f"Server Error: {str(e)}")
        return jsonify({
            'error': 'Server error',
            'message': str(e)
        }), 500

@app.route('/api/download-excel/<filename>', methods=['GET'])
def download_excel(filename):
    try:
        # Construct the full file path
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(filepath):
            return jsonify({
                'error': 'File not found',
                'message': 'The requested file does not exist'
            }), 404
        
        try:
            @after_this_request
            def remove_file(response):
                try:
                    os.remove(filepath)
                except Exception as e:
                    print(f"Error removing file: {e}")
                return response
            
            return send_file(
                filepath,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
            
        except Exception as e:
            if os.path.exists(filepath):
                os.remove(filepath)
            raise
            
    except Exception as e:
        return jsonify({
            'error': 'Download failed',
            'message': str(e)
        }), 500

# Add a test route to verify the server is running
@app.route('/api/test', methods=['GET'])
def test():
    return jsonify({
        "success": True,
        "message": "Server is running"
    })

if __name__ == '__main__':
    print("Starting Financial Advisor Server...")
    print(f"OpenAI API Key configured: {'Yes' if openai.api_key else 'No'}")
    app.run(debug=True, port=5000) 